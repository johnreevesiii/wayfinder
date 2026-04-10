"""
Extract and summarize IHS facility data from the official IHS Excel file.
Source: https://www.ihs.gov/sites/locations/themes/responsive2017/display_objects/documents/ihs_facilities.xlsx
"""
import json
import openpyxl
from pathlib import Path
from collections import Counter

XLSX_PATH = Path(r"C:\Users\johnr\chap\ihs_facilities.xlsx")
OUTPUT_PATH = Path(r"C:\Users\johnr\chap\ihs_facilities_extracted.json")

wb = openpyxl.load_workbook(str(XLSX_PATH), read_only=True)
ws = wb[wb.sheetnames[0]]

rows = list(ws.iter_rows(values_only=True))
wb.close()

# Find the header row (row 3 is the real header: ASUFAC, ASUFAC & MODIFIER, AREA, ...)
header_idx = None
for i, row in enumerate(rows):
    if row and row[0] and str(row[0]).strip() == 'ASUFAC':
        header_idx = i
        break

if header_idx is None:
    print("ERROR: Could not find header row")
    exit(1)

headers = [str(h).strip() if h else f'col_{i}' for i, h in enumerate(rows[header_idx])]
# Only keep meaningful columns (up to COMMENTS, col 28)
headers = headers[:29]
print(f"Header row index: {header_idx}")
print(f"Columns ({len(headers)}):")
for i, h in enumerate(headers):
    print(f"  [{i}] {h}")
print()

# Parse data rows
facilities = []
for row in rows[header_idx + 1:]:
    # Skip empty rows
    if not row or all(v is None for v in row[:5]):
        continue

    vals = list(row[:29])
    # Pad if needed
    while len(vals) < 29:
        vals.append(None)

    rec = {}
    for i, h in enumerate(headers):
        v = vals[i]
        if v is None:
            rec[h] = None
        elif isinstance(v, str) and v.startswith('='):
            # Skip Excel formulas (like =LEFT(B4, 6))
            rec[h] = None
        else:
            rec[h] = v

    # Skip if no facility name
    if not rec.get('FACILITY NAME'):
        continue

    facilities.append(rec)

print(f"Total facility records: {len(facilities)}")
print()

# ---- Statistics ----

# By Area
areas = Counter(f.get('AREA', 'UNKNOWN') for f in facilities)
print("Facilities by IHS Area:")
for area, count in sorted(areas.items(), key=lambda x: -x[1]):
    print(f"  {area}: {count}")
print()

# By Facility Type
ftypes = Counter(f.get('FACILITY TYPE', 'UNKNOWN') for f in facilities)
print("Facilities by Type:")
for ft, count in sorted(ftypes.items(), key=lambda x: -x[1]):
    print(f"  {ft}: {count}")
print()

# By State
states = Counter(f.get('STATE', 'UNKNOWN') for f in facilities)
print("Facilities by State:")
for st, count in sorted(states.items()):
    print(f"  {st}: {count}")
print()

# ITU Code (IHS/Tribal/Urban)
itu = Counter(f.get('ITU CODE', 'UNKNOWN') for f in facilities)
print("ITU Classification:")
for code, count in sorted(itu.items(), key=lambda x: -x[1]):
    print(f"  {code}: {count}")
print()

# Org Type
org = Counter(f.get('ORG TYPE', 'UNKNOWN') for f in facilities)
print("Organization Type:")
for o, count in sorted(org.items(), key=lambda x: -x[1]):
    print(f"  {o}: {count}")
print()

# Field coverage
print("Field coverage:")
for h in headers:
    filled = sum(1 for f in facilities if f.get(h) is not None and str(f.get(h, '')).strip())
    print(f"  {h}: {filled}/{len(facilities)} ({100*filled//len(facilities)}%)")
print()

# Status
status = Counter(f.get('STATUS', 'UNKNOWN') for f in facilities)
print("Status:")
for s, count in sorted(status.items(), key=lambda x: -x[1]):
    print(f"  {s}: {count}")
print()

# Bed counts
beds = []
for f in facilities:
    bc = f.get('BED COUNT')
    if bc is not None:
        try:
            bc = int(bc)
            if bc > 0:
                beds.append(bc)
        except (ValueError, TypeError):
            pass
print(f"Facilities with beds: {len(beds)}")
if beds:
    print(f"  Total beds: {sum(beds)}")
    print(f"  Range: {min(beds)} - {max(beds)}")
print()

# Coordinates coverage
has_coords = sum(1 for f in facilities if f.get('LATITUDE') and f.get('LONGITUDE'))
print(f"With coordinates: {has_coords}/{len(facilities)}")
print()

# ---- Sample records ----
print("=" * 70)
print("SAMPLE RECORDS (first 3):")
print("=" * 70)
for f in facilities[:3]:
    clean = {k: v for k, v in f.items() if v is not None}
    print(json.dumps(clean, indent=2, default=str))
    print()

# ---- Save ----
# Convert all values to JSON-safe types
for f in facilities:
    for k, v in f.items():
        if isinstance(v, (int, float, str, bool)) or v is None:
            continue
        f[k] = str(v)

output = {
    "source": "https://www.ihs.gov/sites/locations/themes/responsive2017/display_objects/documents/ihs_facilities.xlsx",
    "sheet_name": wb.sheetnames[0] if wb.sheetnames else "unknown",
    "extraction_date": "2026-04-10",
    "total_facilities": len(facilities),
    "fields": headers,
    "facilities": facilities,
}

OUTPUT_PATH.write_text(json.dumps(output, indent=2, default=str), encoding="utf-8")
print(f"Saved {len(facilities)} facilities to: {OUTPUT_PATH}")
print(f"Output file size: {OUTPUT_PATH.stat().st_size:,} bytes")
